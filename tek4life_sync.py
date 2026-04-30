"""
tek4life_sync.py — Bee Supplier · iPhones Grade A+
====================================================
Vai buscar iPhones recondicionados Grade A+ ao tek4life.pt,
compara com o snapshot anterior e:

  ✅ Deteta preços que subiram / desceram
  🆕 Alerta de novos produtos
  📦 Deteta mudanças de stock (disponível ↔ esgotado)
  📄 Gera CSV pronto a importar no Bee Supplier (secção Upload)
  🛍  [Opcional] Cria novos produtos como Rascunho no Shopify

Como correr:
    pip install requests openpyxl
    python tek4life_sync.py

Para criar rascunhos automaticamente no Shopify:
    SHOPIFY_TOKEN=shpat_xxx python tek4life_sync.py
    (ou define o token dentro do script em SHOPIFY_TOKEN_MANUAL)
"""

import requests
import csv
import json
import time
import re
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ─── CONFIG ──────────────────────────────────────────────────────────────────

BASE_URL   = "https://www.tek4life.pt"
COLLECTION = "recondicionados"
GRADE_TAGS = ["grade a+", "grade-a+", "a+"]
IPHONE_KW  = "iphone"
LIMIT      = 250
HEADERS    = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
}

# Shopify (Bee Supplier)
SHOPIFY_PROXY   = "https://shopify-proxy.victoraurich11.workers.dev"
SHOPIFY_SHOP    = "bee-store-loja.myshopify.com"
SHOPIFY_VERSION = "2026-01"
SHOPIFY_TOKEN_MANUAL = ""  # ← preenche aqui se não usares variável de ambiente

# Stock virtual atribuído quando tek4life expõe `available=true`.
# A products.json pública não revela inventory_quantity real, por isso usamos
# um valor fixo > threshold de "low stock" da theme BeeStore (10) para evitar
# o badge "Apenas N unidade(s)" em todos os recondicionados.
STOCK_VIRTUAL_DISPONIVEL = 15

# Ficheiros de output
# IMPORTANTE: launchd em macOS NÃO tem acesso a ~/Downloads (sandbox de privacidade
# desde Catalina). Por isso escrevemos PRIMEIRO no SCRIPT_DIR (sempre acessível) e
# depois tentamos copiar para Downloads — se a cópia falhar o launchd não crasha.
SCRIPT_DIR    = Path(__file__).parent
SNAPSHOT_FILE = SCRIPT_DIR / "tek4life_snapshot.json"       # snapshot fica junto ao script
LAST_RUN_FILE = SCRIPT_DIR / "tek4life_last_run.json"       # registo da última corrida
DOWNLOADS_DIR = Path.home() / "Downloads"
OUTPUT_CSV    = SCRIPT_DIR / "tek4life_upload.csv"          # CSV: primário no script dir
OUTPUT_REPORT = SCRIPT_DIR / "tek4life_relatorio.xlsx"      # Relatório: idem
OUTPUT_CSV_COPY    = DOWNLOADS_DIR / "tek4life_upload.csv"      # cópia best-effort
OUTPUT_REPORT_COPY = DOWNLOADS_DIR / "tek4life_relatorio.xlsx"  # cópia best-effort

def try_copy_to_downloads(src: Path, dst: Path) -> bool:
    """Tenta copiar para Downloads. Se falhar (launchd sem permissão), regista warning e continua."""
    try:
        import shutil
        shutil.copy2(src, dst)
        print(f"   📋 Cópia em Downloads: {dst}")
        return True
    except (PermissionError, OSError) as e:
        print(f"   ⚠️  Não foi possível copiar para Downloads ({e.__class__.__name__}: {e}). Ficheiro disponível em: {src}")
        return False

# ─── CONTROLO DE EXECUÇÃO DIÁRIA ─────────────────────────────────────────────

def already_ran_today():
    """Devolve True se o script já correu hoje (evita duplicados se o Mac ligar tarde)."""
    if not LAST_RUN_FILE.exists():
        return False
    try:
        data = json.loads(LAST_RUN_FILE.read_text())
        return data.get("date") == datetime.now().strftime("%Y-%m-%d")
    except Exception:
        return False

def mark_ran_today():
    LAST_RUN_FILE.write_text(json.dumps({
        "date": datetime.now().strftime("%Y-%m-%d"),
        "time": datetime.now().strftime("%H:%M:%S")
    }))

# ─── FETCH TEK4LIFE ──────────────────────────────────────────────────────────

def fetch_all_products():
    products, page = [], 1
    print(f"🔍 A carregar produtos do tek4life...")
    while True:
        url = f"{BASE_URL}/collections/{COLLECTION}/products.json?limit={LIMIT}&page={page}"
        try:
            r = requests.get(url, headers=HEADERS, timeout=30)
            r.raise_for_status()
            data = r.json().get("products", [])
        except Exception as e:
            print(f"  ❌ Erro na página {page}: {e}")
            break
        if not data:
            break
        products.extend(data)
        print(f"  Página {page}: {len(data)} produtos")
        if len(data) < LIMIT:
            break
        page += 1
        time.sleep(0.4)
    print(f"  Total: {len(products)} produtos\n")
    return products


def is_iphone(p):
    t = (p.get("title") or "").lower()
    pt = (p.get("product_type") or "").lower()
    v = (p.get("vendor") or "").lower()
    tags = [x.lower() for x in p.get("tags", [])]
    return IPHONE_KW in t or IPHONE_KW in pt or "apple" in v or any(IPHONE_KW in x for x in tags)


def has_grade_a_plus(product, variant=None):
    tags = [x.lower() for x in product.get("tags", [])]
    if any(g in tags for g in GRADE_TAGS):
        return True
    if any(g in (product.get("title") or "").lower() for g in GRADE_TAGS):
        return True
    if variant:
        parts = [
            (variant.get("title") or "").lower(),
            *[(variant.get(f"option{i}") or "").lower() for i in range(1, 4)]
        ]
        if any(g in " ".join(parts) for g in GRADE_TAGS):
            return True
    return False


def strip_html(html):
    if not html:
        return ""
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html)).strip()


def parse_iphones(all_products):
    """Retorna dict keyed by SKU → dados normalizados."""
    result = {}
    for p in all_products:
        if not is_iphone(p):
            continue
        prod_grade = has_grade_a_plus(p)
        for v in p.get("variants", []):
            if not prod_grade and not has_grade_a_plus(p, v):
                continue
            sku = (v.get("sku") or "").strip()
            if not sku:
                # fallback: usa handle + opções como chave única
                opts = "_".join(filter(None, [v.get(f"option{i}") for i in range(1, 4)]))
                sku = f"{p['handle']}_{opts}"
            vt = v.get("title", "")
            full_title = p["title"]
            if vt and vt.lower() != "default title":
                full_title = f"{p['title']} — {vt}"
            images = p.get("images", [])
            result[sku] = {
                "title":         full_title,
                "base_title":    p["title"],
                "sku":           sku,
                "price":         float(v.get("price") or 0),
                "compare_price": float(v.get("compare_at_price") or 0) if v.get("compare_at_price") else None,
                "available":     bool(v.get("available", False)),
                "stock":         v.get("inventory_quantity", 0) or 0,
                "option1":       v.get("option1") or "",
                "option2":       v.get("option2") or "",
                "option3":       v.get("option3") or "",
                "image":         images[0]["src"] if images else "",
                "brand":         p.get("vendor") or "Apple",
                "description":   strip_html(p.get("body_html") or ""),
                "url":           f"{BASE_URL}/products/{p['handle']}",
                "handle":        p["handle"],
                "fetched_at":    datetime.now().isoformat(),
            }
    return result


# ─── SNAPSHOT ────────────────────────────────────────────────────────────────

def load_snapshot():
    if SNAPSHOT_FILE.exists():
        with open(SNAPSHOT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_snapshot(data):
    with open(SNAPSHOT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"💾 Snapshot guardado: {SNAPSHOT_FILE.name}")


# ─── COMPARAÇÃO ──────────────────────────────────────────────────────────────

def compare(current, previous):
    """
    Retorna dicionário com listas de mudanças:
      new_products   → SKUs que não existiam antes
      price_up       → preço subiu
      price_down     → preço baixou
      back_in_stock  → voltou a estar disponível
      out_of_stock   → ficou esgotado
      unchanged      → sem alterações
    """
    changes = {
        "new_products":  [],
        "price_up":      [],
        "price_down":    [],
        "back_in_stock": [],
        "out_of_stock":  [],
        "unchanged":     [],
    }

    for sku, curr in current.items():
        if sku not in previous:
            changes["new_products"].append(curr)
            continue

        prev = previous[sku]
        price_changed = abs(curr["price"] - prev["price"]) >= 0.01
        avail_changed = curr["available"] != prev["available"]

        if price_changed:
            entry = {**curr, "prev_price": prev["price"]}
            if curr["price"] > prev["price"]:
                changes["price_up"].append(entry)
            else:
                changes["price_down"].append(entry)

        if avail_changed:
            if curr["available"] and not prev["available"]:
                changes["back_in_stock"].append(curr)
            elif not curr["available"] and prev["available"]:
                changes["out_of_stock"].append(curr)

        if not price_changed and not avail_changed:
            changes["unchanged"].append(curr)

    return changes


# ─── PRINT REPORT ────────────────────────────────────────────────────────────

def print_report(changes, current, previous):
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    prev_date = "—"
    if previous:
        # Tenta obter a data do snapshot anterior
        sample = next(iter(previous.values()), {})
        if "fetched_at" in sample:
            try:
                prev_date = datetime.fromisoformat(sample["fetched_at"]).strftime("%d/%m/%Y %H:%M")
            except Exception:
                pass

    print("=" * 62)
    print(f"  🐝 Bee Supplier — Sync tek4life")
    print(f"  📅 Agora: {now}   |   Snapshot anterior: {prev_date}")
    print(f"  📱 iPhones Grade A+: {len(current)} variantes")
    print("=" * 62)

    if not previous:
        print("\n  ℹ  Primeiro sync — sem histórico para comparar.")
        print(f"  Todos os {len(current)} produtos foram guardados no snapshot.\n")
        return

    if changes["new_products"]:
        print(f"\n🆕 NOVOS PRODUTOS ({len(changes['new_products'])})")
        print("─" * 50)
        for p in changes["new_products"]:
            avail = "✅ Disponível" if p["available"] else "❌ Esgotado"
            print(f"  + {p['title']}")
            print(f"    SKU: {p['sku']} | €{p['price']:.2f} | {avail}")

    if changes["price_down"]:
        print(f"\n📉 PREÇO BAIXOU ({len(changes['price_down'])})")
        print("─" * 50)
        for p in changes["price_down"]:
            diff = p["prev_price"] - p["price"]
            print(f"  ↓ {p['title']}")
            print(f"    €{p['prev_price']:.2f} → €{p['price']:.2f}  (-€{diff:.2f})")

    if changes["price_up"]:
        print(f"\n📈 PREÇO SUBIU ({len(changes['price_up'])})")
        print("─" * 50)
        for p in changes["price_up"]:
            diff = p["price"] - p["prev_price"]
            print(f"  ↑ {p['title']}")
            print(f"    €{p['prev_price']:.2f} → €{p['price']:.2f}  (+€{diff:.2f})")

    if changes["back_in_stock"]:
        print(f"\n♻️  VOLTOU AO STOCK ({len(changes['back_in_stock'])})")
        print("─" * 50)
        for p in changes["back_in_stock"]:
            print(f"  ✅ {p['title']}  (SKU: {p['sku']})")

    if changes["out_of_stock"]:
        print(f"\n🔴 FICOU ESGOTADO ({len(changes['out_of_stock'])})")
        print("─" * 50)
        for p in changes["out_of_stock"]:
            print(f"  ❌ {p['title']}  (SKU: {p['sku']})")

    total_changes = sum(len(v) for k, v in changes.items() if k != "unchanged")
    if total_changes == 0:
        print("\n✅ Sem alterações desde o último sync.")
    else:
        print(f"\n📊 Resumo: {len(changes['new_products'])} novos | "
              f"{len(changes['price_down'])} preço↓ | {len(changes['price_up'])} preço↑ | "
              f"{len(changes['back_in_stock'])} voltou stock | {len(changes['out_of_stock'])} esgotou | "
              f"{len(changes['unchanged'])} sem alteração")
    print()


# ─── CSV PARA BEE SUPPLIER ───────────────────────────────────────────────────
# Colunas compatíveis com o mapeador automático do Bee Supplier

def export_csv(current):
    """
    Gera CSV com colunas que o Bee Supplier auto-deteta no Upload.
    Campos: Title, SKU, Price, Stock, Image, Brand
    """
    rows = sorted(current.values(), key=lambda x: x["title"])
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["Title", "SKU", "Price", "Stock", "Image", "Brand"])
        writer.writeheader()
        for p in rows:
            writer.writerow({
                "Title":  p["title"],
                "SKU":    p["sku"],
                "Price":  f"{p['price']:.2f}",
                "Stock":  STOCK_VIRTUAL_DISPONIVEL if p["available"] else 0,
                "Image":  p["image"],
                "Brand":  p["brand"],
            })
    print(f"📄 CSV Bee Supplier: {OUTPUT_CSV}  ({len(rows)} linhas)")
    try_copy_to_downloads(OUTPUT_CSV, OUTPUT_CSV_COPY)


# ─── EXCEL RELATÓRIO ─────────────────────────────────────────────────────────

def export_excel_report(changes, current, previous):
    if not EXCEL_OK:
        print("⚠️  openpyxl não disponível — relatório Excel ignorado.")
        return

    wb = openpyxl.Workbook()

    # Paleta BeeStore
    COLOR_YELLOW = "F5C400"
    COLOR_DARK   = "1A1F2E"
    COLOR_GREEN  = "16A34A"
    COLOR_RED    = "DC2626"
    COLOR_BLUE   = "1B6FE8"
    COLOR_ORANGE = "D97706"
    COLOR_LIGHT  = "F4F6FA"

    def hfont(bold=True, color="FFFFFF", size=11):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color="E2E6EF")
        return Border(left=s, right=s, top=s, bottom=s)

    def make_sheet(title, color_hex, rows_data, columns, col_widths):
        ws = wb.create_sheet(title=title)
        ws.sheet_properties.tabColor = color_hex

        # Header
        for col_i, col_name in enumerate(columns, 1):
            c = ws.cell(row=1, column=col_i, value=col_name)
            c.fill = fill(COLOR_DARK)
            c.font = hfont()
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
        ws.row_dimensions[1].height = 22

        # Rows
        for row_i, row in enumerate(rows_data, 2):
            bg = COLOR_LIGHT if row_i % 2 == 0 else "FFFFFF"
            for col_i, val in enumerate(row, 1):
                c = ws.cell(row=row_i, column=col_i, value=val)
                c.fill = fill(bg)
                c.border = thin_border()
                c.alignment = Alignment(vertical="center", wrap_text=False)

        # Column widths
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        return ws

    # ── Folha: Resumo ────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "📊 Resumo"
    ws_sum.sheet_properties.tabColor = COLOR_YELLOW

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    summary_data = [
        ("Data do sync", now_str),
        ("Total iPhones A+", len(current)),
        ("🆕 Novos produtos", len(changes["new_products"])),
        ("📉 Preço baixou", len(changes["price_down"])),
        ("📈 Preço subiu", len(changes["price_up"])),
        ("♻️  Voltou ao stock", len(changes["back_in_stock"])),
        ("🔴 Ficou esgotado", len(changes["out_of_stock"])),
        ("✅ Sem alterações", len(changes["unchanged"])),
    ]
    row_colors = {
        "🆕 Novos produtos":  ("DBEAFE", COLOR_BLUE),
        "📉 Preço baixou":    ("DCFCE7", COLOR_GREEN),
        "📈 Preço subiu":     ("FEF3C7", COLOR_ORANGE),
        "♻️  Voltou ao stock": ("DCFCE7", COLOR_GREEN),
        "🔴 Ficou esgotado":  ("FEF2F2", COLOR_RED),
    }
    ws_sum["A1"] = "🐝 Bee Supplier — Sync tek4life"
    ws_sum["A1"].font = Font(bold=True, size=16, color=COLOR_DARK, name="Calibri")
    ws_sum.merge_cells("A1:B1")
    ws_sum.row_dimensions[1].height = 32

    for i, (label, val) in enumerate(summary_data, 2):
        bg, fg = row_colors.get(label, ("F4F6FA", COLOR_DARK))
        lc = ws_sum.cell(row=i, column=1, value=label)
        vc = ws_sum.cell(row=i, column=2, value=val)
        for c in [lc, vc]:
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center")
        lc.font = Font(bold=True, color=COLOR_DARK, name="Calibri")
        vc.font = Font(bold=True, color=fg, size=13, name="Calibri")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[i].height = 20

    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 14

    # ── Folha: Novos Produtos ────────────────────────────────────────────────
    if changes["new_products"]:
        rows_new = [
            (p["title"], p["sku"], f"€{p['price']:.2f}",
             "✅ Sim" if p["available"] else "❌ Não",
             p["stock"], p["url"])
            for p in sorted(changes["new_products"], key=lambda x: x["title"])
        ]
        ws_new = make_sheet("🆕 Novos", COLOR_BLUE,
            rows_new,
            ["Produto", "SKU", "Preço", "Disponível", "Stock", "URL"],
            [45, 22, 12, 14, 10, 50])

        # Pintar linha inteira a azul claro
        for row_i in range(2, len(rows_new) + 2):
            for col_i in range(1, 7):
                ws_new.cell(row_i, col_i).fill = fill("DBEAFE")

    # ── Folha: Preço Baixou ──────────────────────────────────────────────────
    if changes["price_down"]:
        rows_down = [
            (p["title"], p["sku"],
             f"€{p['prev_price']:.2f}", f"€{p['price']:.2f}",
             f"-€{p['prev_price'] - p['price']:.2f}",
             f"{((p['prev_price'] - p['price']) / p['prev_price'] * 100):.1f}%")
            for p in sorted(changes["price_down"], key=lambda x: x["prev_price"] - x["price"], reverse=True)
        ]
        make_sheet("📉 Preço Baixou", COLOR_GREEN,
            rows_down,
            ["Produto", "SKU", "Preço Anterior", "Preço Atual", "Diferença", "% Descida"],
            [45, 22, 16, 14, 13, 12])

    # ── Folha: Preço Subiu ───────────────────────────────────────────────────
    if changes["price_up"]:
        rows_up = [
            (p["title"], p["sku"],
             f"€{p['prev_price']:.2f}", f"€{p['price']:.2f}",
             f"+€{p['price'] - p['prev_price']:.2f}",
             f"+{((p['price'] - p['prev_price']) / p['prev_price'] * 100):.1f}%")
            for p in sorted(changes["price_up"], key=lambda x: x["price"] - x["prev_price"], reverse=True)
        ]
        ws_up = make_sheet("📈 Preço Subiu", COLOR_ORANGE,
            rows_up,
            ["Produto", "SKU", "Preço Anterior", "Preço Atual", "Diferença", "% Subida"],
            [45, 22, 16, 14, 13, 12])
        # Pintar preço atual a laranja
        for row_i in range(2, len(rows_up) + 2):
            ws_up.cell(row_i, 4).font = Font(bold=True, color=COLOR_ORANGE, name="Calibri")

    # ── Folha: Stock ─────────────────────────────────────────────────────────
    stock_changes = [
        (*([p["title"], p["sku"], "♻️ Voltou"]), ) for p in changes["back_in_stock"]
    ] + [
        (*([p["title"], p["sku"], "🔴 Esgotou"]), ) for p in changes["out_of_stock"]
    ]
    if stock_changes:
        make_sheet("📦 Stock", COLOR_RED,
            stock_changes,
            ["Produto", "SKU", "Mudança"],
            [45, 22, 16])

    # ── Folha: Catálogo Completo ─────────────────────────────────────────────
    all_rows = []
    for p in sorted(current.values(), key=lambda x: x["title"]):
        sku = p["sku"]
        if sku in {x["sku"] for x in changes["new_products"]}:
            status = "🆕 Novo"
        elif sku in {x["sku"] for x in changes["price_down"]}:
            status = "📉 Preço↓"
        elif sku in {x["sku"] for x in changes["price_up"]}:
            status = "📈 Preço↑"
        elif sku in {x["sku"] for x in changes["out_of_stock"]}:
            status = "🔴 Esgotou"
        elif sku in {x["sku"] for x in changes["back_in_stock"]}:
            status = "♻️ Voltou"
        else:
            status = "✅ Sem alteração"

        prev_price = ""
        if sku in previous:
            prev_price = f"€{previous[sku]['price']:.2f}"

        all_rows.append((
            p["title"], p["sku"],
            f"€{p['price']:.2f}", prev_price,
            "✅" if p["available"] else "❌",
            p["stock"], status
        ))

    make_sheet("📋 Catálogo", "888888",
        all_rows,
        ["Produto", "SKU", "Preço Atual", "Preço Anterior", "Disponível", "Stock", "Status"],
        [45, 22, 14, 16, 12, 10, 18])

    wb.save(OUTPUT_REPORT)
    print(f"📊 Relatório Excel: {OUTPUT_REPORT}")
    try_copy_to_downloads(OUTPUT_REPORT, OUTPUT_REPORT_COPY)


# ─── SHOPIFY: CRIAR RASCUNHO ─────────────────────────────────────────────────

def get_shopify_token():
    token = os.environ.get("SHOPIFY_TOKEN") or SHOPIFY_TOKEN_MANUAL
    if not token:
        token = input("\n🔑 Token Shopify (deixa em branco para saltar): ").strip()
    return token


def shopify_gql(query, variables, token):
    resp = requests.post(
        SHOPIFY_PROXY,
        headers={
            "Content-Type":      "application/json",
            "X-Shopify-Shop":    SHOPIFY_SHOP,
            "X-Shopify-Token":   token,
            "X-Shopify-Version": SHOPIFY_VERSION,
        },
        json={"query": query, "variables": variables},
        timeout=30,
    )
    resp.raise_for_status()
    data = resp.json()
    if data.get("errors"):
        raise Exception(data["errors"][0]["message"])
    return data.get("data", {})


def create_draft_product(p, token):
    """
    Cria produto como DRAFT no Shopify via productCreate.
    Inclui preço, stock e imagem se disponível.
    """
    mutation = """
    mutation($input: ProductInput!, $media: [CreateMediaInput!]) {
      productCreate(input: $input, media: $media) {
        product {
          id
          variants(first: 1) { edges { node { id inventoryItem { id } } } }
        }
        userErrors { message }
      }
    }
    """
    variables = {
        "input": {
            "title":  p["title"],
            "vendor": p["brand"] or "Apple",
            "status": "DRAFT",
            "tags":   ["grade a+", "recondicionado", "iphone", "tek4life", "cat:sem-origem"],
        },
        "media": [
            {"originalSource": p["image"], "mediaContentType": "IMAGE", "alt": p["title"]}
        ] if p.get("image") else None,
    }
    data = shopify_gql(mutation, variables, token)
    result = data.get("productCreate", {})
    if result.get("userErrors"):
        raise Exception(result["userErrors"][0]["message"])

    product = result["product"]
    product_id = product["id"]
    variant_node = product["variants"]["edges"][0]["node"]
    variant_id = variant_node["id"]
    inv_item_id = variant_node["inventoryItem"]["id"]

    # Atualiza preço
    price_mutation = """
    mutation($productId: ID!, $variants: [ProductVariantsBulkInput!]!) {
      productVariantsBulkUpdate(productId: $productId, variants: $variants) {
        productVariants { id price }
        userErrors { message }
      }
    }
    """
    shopify_gql(price_mutation, {
        "productId": product_id,
        "variants": [{"id": variant_id, "price": str(p["price"])}]
    }, token)

    return product_id


def create_drafts(new_products, token):
    """Cria todos os novos produtos como rascunho no Shopify."""
    if not new_products:
        return
    created, errors = 0, 0
    print(f"\n🛍  A criar {len(new_products)} rascunho(s) no Shopify...")
    for p in new_products:
        try:
            pid = create_draft_product(p, token)
            print(f"  ✅ {p['title']}")
            print(f"     → Shopify ID: {pid}")
            created += 1
        except Exception as e:
            print(f"  ❌ {p['title']}: {e}")
            errors += 1
        time.sleep(0.5)  # evitar rate limit
    print(f"\n  Criados: {created} | Erros: {errors}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "=" * 62)
    print("  🐝 Bee Supplier — Sync tek4life · iPhones Grade A+")
    print(f"  {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 62 + "\n")

    # Verifica se já correu hoje (catch-up após Mac ligar tarde)
    if already_ran_today():
        print("✅ Sync já realizado hoje. A sair.")
        return

    # 1. Fetch
    all_products = fetch_all_products()
    if not all_products:
        print("❌ Não foi possível carregar produtos. Verifica a ligação.")
        return

    current = parse_iphones(all_products)
    if not current:
        print("⚠️  Nenhum iPhone Grade A+ encontrado.")
        return

    # 2. Comparar com snapshot anterior
    previous = load_snapshot()
    changes = compare(current, previous)

    # 3. Relatório no terminal
    print_report(changes, current, previous)

    # 4. Guardar snapshot atualizado
    save_snapshot(current)

    # 5. CSV para Bee Supplier (Upload)
    export_csv(current)

    # 6. Relatório Excel com todas as mudanças
    export_excel_report(changes, current, previous)

    # 7. Novos produtos — criar rascunhos directo no Bee Supplier (secção Tek4life › Sem match)
    if changes["new_products"]:
        print(f"\n🆕 {len(changes['new_products'])} produto(s) novo(s) detetado(s).")
        print(f"   → Abre o Bee Supplier › Tek4life › filtro 'Sem match' para criar rascunhos.")

    # Só agora, depois de todos os outputs terem corrido sem crashes, registamos o "ran today".
    # Se algo crashar antes disto (ex: PermissionError no export), as próximas execuções de
    # catch-up (de hora a hora até às 13h) vão tentar novamente em vez de saltar.
    mark_ran_today()

    print(f"\n✅ Sync concluído!")
    print(f"   → {OUTPUT_CSV.name}  (importa no Bee Supplier › Upload)")
    if EXCEL_OK:
        print(f"   → {OUTPUT_REPORT.name}  (relatório completo de mudanças)")
    print()


if __name__ == "__main__":
    main()
